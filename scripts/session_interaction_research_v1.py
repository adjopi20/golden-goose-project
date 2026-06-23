from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from typing import Literal
import numpy as np
import pandas as pd
from openpyxl.styles import numbers


WIB_TZ = "Asia/Jakarta"
UTC_TZ = "UTC"


# =============================================================================
# General helpers
# =============================================================================


def safe_divide(a, b):
    """Return a / b, with division-by-zero mapped to NaN."""
    if np.isscalar(a) and np.isscalar(b):
        if pd.isna(a) or pd.isna(b) or b == 0:
            return np.nan
        return a / b

    a_arr = np.asarray(a, dtype="float64")
    b_arr = np.asarray(b, dtype="float64")
    with np.errstate(divide="ignore", invalid="ignore"):
        out = np.divide(
            a_arr,
            b_arr,
            out=np.full_like(a_arr, np.nan, dtype="float64"),
            where=(~np.isnan(a_arr)) & (~np.isnan(b_arr)) & (b_arr != 0),
        )
    return out


def pick_col(columns, candidates, required=True):
    cols_lower = {c.lower(): c for c in columns}
    for cand in candidates:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    if required:
        raise ValueError(f"Could not find required column. Tried: {candidates}. Available: {list(columns)}")
    return None


def normalize_timestamp(series: pd.Series) -> pd.Series:
    """Normalize various numeric timestamp units to timezone-aware UTC."""
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, utc=True)

    x = pd.to_numeric(series, errors="coerce")
    if x.notna().sum() == 0:
        return pd.to_datetime(x, utc=True)

    max_abs = float(np.nanmax(np.abs(x.to_numpy(dtype="float64"))))
    if max_abs > 1e17:
        unit = "ns"
    elif max_abs > 1e14:
        unit = "us"
    elif max_abs > 1e11:
        unit = "ms"
    else:
        unit = "s"

    return pd.to_datetime(x, unit=unit, utc=True)


def parse_hhmm(value: str) -> tuple[int, int]:
    try:
        h, m = value.split(":")
        h_i = int(h)
        m_i = int(m)
    except Exception as exc:
        raise ValueError(f"Invalid HH:MM value: {value}") from exc
    if not (0 <= h_i <= 23 and 0 <= m_i <= 59):
        raise ValueError(f"Invalid HH:MM value: {value}")
    return h_i, m_i


def format_ts_utc(ts) -> str | None:
    if pd.isna(ts):
        return None
    ts = pd.Timestamp(ts)
    if ts.tzinfo is None:
        ts = ts.tz_localize(UTC_TZ)
    else:
        ts = ts.tz_convert(UTC_TZ)
    return ts.strftime("%Y-%m-%d %H:%M:%S UTC")


def format_ts_wib(ts) -> str | None:
    if pd.isna(ts):
        return None
    ts = pd.Timestamp(ts)
    if ts.tzinfo is None:
        ts = ts.tz_localize(UTC_TZ)
    ts = ts.tz_convert(WIB_TZ)
    return ts.strftime("%Y-%m-%d %H:%M:%S WIB")


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


# =============================================================================
# Config
# =============================================================================


@dataclass(frozen=True)
class SessionWindow:
    name: str
    start_hour: int
    start_minute: int
    end_hour: int
    end_minute: int
    start_offset_days: int
    end_offset_days: int

    def start_offset(self) -> pd.Timedelta:
        return pd.Timedelta(days=self.start_offset_days, hours=self.start_hour, minutes=self.start_minute)

    def end_offset(self) -> pd.Timedelta:
        return pd.Timedelta(days=self.end_offset_days, hours=self.end_hour, minutes=self.end_minute)


def validate_session_defaults(args) -> None:
    expected = {
        (args.asia_start, args.asia_end): ("23:00", "07:00"),
        (args.europe_start, args.europe_end): ("07:00", "15:00"),
        (args.us_start, args.us_end): ("15:00", "23:00"),
    }
    for actual, wanted in expected.items():
        parse_hhmm(actual[0])
        parse_hhmm(actual[1])
        parse_hhmm(wanted[0])
        parse_hhmm(wanted[1])


def build_session_windows(args) -> list[SessionWindow]:
    """
    V1 implementation targets the requested fixed 24-hour framework:
    Asia 23:00-07:00 UTC, Europe 07:00-15:00 UTC, US 15:00-23:00 UTC.
    session_day_wib is the WIB date of Asia session start / Europe UTC date anchor.
    """
    a_sh, a_sm = parse_hhmm(args.asia_start)
    a_eh, a_em = parse_hhmm(args.asia_end)
    e_sh, e_sm = parse_hhmm(args.europe_start)
    e_eh, e_em = parse_hhmm(args.europe_end)
    u_sh, u_sm = parse_hhmm(args.us_start)
    u_eh, u_em = parse_hhmm(args.us_end)

    windows = [
        SessionWindow("asia", a_sh, a_sm, a_eh, a_em, -1, 0),
        SessionWindow("europe", e_sh, e_sm, e_eh, e_em, 0, 0),
        SessionWindow("us", u_sh, u_sm, u_eh, u_em, 0, 0),
    ]

    # Validate intended continuity and 8-hour windows.
    anchor = pd.Timestamp("2026-01-02 00:00:00", tz=UTC_TZ)
    prev_end = None
    for win in windows:
        start = anchor + win.start_offset()
        end = anchor + win.end_offset()
        if end <= start:
            raise ValueError(f"Session {win.name} must end after start in anchored construction.")
        hours = (end - start).total_seconds() / 3600
        if abs(hours - 8.0) > 1e-9:
            raise ValueError(f"Session {win.name} must be 8 hours. Got {hours}.")
        if prev_end is not None and start != prev_end:
            raise ValueError("Sessions must be contiguous in the 24-hour framework.")
        prev_end = end
    return windows


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Asia/Europe/US session interaction research on raw Binance aggTrade parquet.")
    parser.add_argument("--input", required=True, help="Input raw aggTrade parquet path")
    parser.add_argument("--symbol", required=True, help="Symbol name, e.g. AVAXUSDC")
    parser.add_argument("--output-xlsx", required=True, help="Output Excel workbook path")

    parser.add_argument("--asia-start", default="23:00")
    parser.add_argument("--asia-end", default="07:00")
    parser.add_argument("--europe-start", default="07:00")
    parser.add_argument("--europe-end", default="15:00")
    parser.add_argument("--us-start", default="15:00")
    parser.add_argument("--us-end", default="23:00")

    parser.add_argument("--direction-eff-threshold", type=float, default=0.25)
    parser.add_argument("--low-body-threshold", type=float, default=0.25)
    parser.add_argument("--high-body-threshold", type=float, default=0.60)
    parser.add_argument("--range-low-quantile", type=float, default=0.33)
    parser.add_argument("--range-high-quantile", type=float, default=0.67)
    parser.add_argument("--sweep-break-share-threshold", type=float, default=0.10)
    parser.add_argument("--extreme-close-threshold", type=float, default=0.10)
    parser.add_argument("--similar-range-low", type=float, default=0.50)
    parser.add_argument("--similar-range-high", type=float, default=1.50)
    parser.add_argument("--min-trades-per-session", type=int, default=10)
    parser.add_argument("--random-seed", type=int, default=42)
    return parser


def validate_args(args) -> None:
    if args.input is None or args.symbol is None or args.output_xlsx is None:
        raise ValueError("--input, --symbol, and --output-xlsx are required.")
    if args.min_trades_per_session <= 0:
        raise ValueError("--min-trades-per-session must be > 0")
    quantiles = [args.range_low_quantile, args.range_high_quantile]
    if not (0 <= quantiles[0] <= 1 and 0 <= quantiles[1] <= 1 and quantiles[0] <= quantiles[1]):
        raise ValueError("Range quantiles must satisfy 0 <= low <= high <= 1")
    for name in [
        "direction_eff_threshold",
        "low_body_threshold",
        "high_body_threshold",
        "sweep_break_share_threshold",
        "extreme_close_threshold",
        "similar_range_low",
        "similar_range_high",
    ]:
        if getattr(args, name) < 0:
            raise ValueError(f"--{name.replace('_', '-')} must be >= 0")
    validate_session_defaults(args)


# =============================================================================
# Data loading and normalization
# =============================================================================


def load_raw_aggtrades(path: Path) -> pd.DataFrame:
    print(f"Loading parquet: {path}")
    if not path.exists():
        raise FileNotFoundError(f"Input parquet not found: {path}")

    import pyarrow.parquet as pq

    pf = pq.ParquetFile(path)
    columns = pf.schema.names

    ts_col = pick_col(columns, ["timestamp", "event_timestamp", "T", "time", "transact_time"])
    price_col = pick_col(columns, ["price", "p"])
    qty_col = pick_col(columns, ["qty", "quantity", "q"])
    maker_col = pick_col(columns, ["is_buyer_maker", "m"], required=False)

    read_cols = [ts_col, price_col, qty_col]
    if maker_col:
        read_cols.append(maker_col)

    df = pd.read_parquet(path, columns=read_cols)
    df = df.rename(columns={ts_col: "timestamp", price_col: "price", qty_col: "qty"})

    if maker_col:
        df = df.rename(columns={maker_col: "is_buyer_maker"})
    else:
        df["is_buyer_maker"] = np.nan

    df["timestamp"] = normalize_timestamp(df["timestamp"])
    df["price"] = pd.to_numeric(df["price"], errors="coerce").astype("float64")
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").astype("float64")
    df = df.dropna(subset=["timestamp", "price", "qty"]).copy()
    df = df.sort_values("timestamp").reset_index(drop=True)
    df["notional"] = df["price"] * df["qty"]

    if df["is_buyer_maker"].notna().any():
        df["is_buyer_maker"] = df["is_buyer_maker"].astype(bool)
        df["aggressor_side"] = np.where(df["is_buyer_maker"], -1, 1)
    else:
        df["aggressor_side"] = np.nan

    if df.empty:
        raise ValueError("Input parquet is empty after cleaning required fields.")

    print(f"Loaded rows: {len(df):,}")
    print(f"Timestamp range UTC: {df['timestamp'].min()} -> {df['timestamp'].max()}")
    return df


# =============================================================================
# Session assignment and aggregation
# =============================================================================


def add_session_day_wib(df: pd.DataFrame) -> pd.DataFrame:
    """
    Map each trade into a WIB trading day anchor.

    Since the 24h framework is:
      Asia   UTC 23:00 prev day -> 07:00 anchor day
      Europe UTC 07:00 -> 15:00 anchor day
      US     UTC 15:00 -> 23:00 anchor day

    the anchor UTC date is:
      - same UTC date for timestamps >= 07:00
      - next UTC date for timestamps in [23:00, 24:00)
    """
    ts = df["timestamp"]
    hour_minute = ts.dt.hour * 60 + ts.dt.minute
    utc_date = ts.dt.floor("D")
    anchor_utc_date = utc_date.where(hour_minute >= 7 * 60, utc_date)
    anchor_utc_date = anchor_utc_date.where(hour_minute < 23 * 60, utc_date + pd.Timedelta(days=1))
    df = df.copy()
    df["anchor_utc_date"] = anchor_utc_date
    df["session_day_wib"] = anchor_utc_date.dt.tz_convert(WIB_TZ).dt.strftime("%Y-%m-%d")
    return df


def build_daily_windows(anchor_utc_date: pd.Timestamp, windows: list[SessionWindow]) -> dict[str, tuple[pd.Timestamp, pd.Timestamp]]:
    out = {}
    for win in windows:
        start = anchor_utc_date + win.start_offset()
        end = anchor_utc_date + win.end_offset()
        out[win.name] = (start, end)
    return out


def aggregate_session_metrics(session_df: pd.DataFrame, prefix: str) -> dict:
    session_df = session_df.sort_values("timestamp")
    open_ = float(session_df["price"].iloc[0])
    high_ = float(session_df["price"].max())
    low_ = float(session_df["price"].min())
    close_ = float(session_df["price"].iloc[-1])
    range_ = high_ - low_
    body = close_ - open_
    abs_body = abs(body)

    metrics = {
        f"{prefix}_open": open_,
        f"{prefix}_high": high_,
        f"{prefix}_low": low_,
        f"{prefix}_close": close_,
        f"{prefix}_range": range_,
        f"{prefix}_range_pct": safe_divide(range_, open_),
        f"{prefix}_body": body,
        f"{prefix}_abs_body": abs_body,
        f"{prefix}_body_pct": safe_divide(body, open_),
        f"{prefix}_body_to_range": safe_divide(abs_body, range_),
        f"{prefix}_efficiency": safe_divide(body, range_),
        f"{prefix}_close_position": safe_divide(close_ - low_, range_),
        f"{prefix}_volume": float(session_df["qty"].sum()),
        f"{prefix}_notional": float(session_df["notional"].sum()),
        f"{prefix}_trade_count": int(len(session_df)),
        f"{prefix}_first_trade_timestamp": session_df["timestamp"].iloc[0],
        f"{prefix}_last_trade_timestamp": session_df["timestamp"].iloc[-1],
    }
    return metrics



def first_break_timestamp(
    session_df: pd.DataFrame,
    high_level: float,
    low_level: float,
    high_cmp: Literal["gt", "gte"],
    low_cmp: Literal["lt", "lte"],
) -> tuple[pd.Timestamp | None, pd.Timestamp | None, str]:
    prices = session_df["price"]

    if high_cmp == "gt":
        high_hits = session_df.loc[prices > high_level, "timestamp"]
    else:
        high_hits = session_df.loc[prices >= high_level, "timestamp"]

    if low_cmp == "lt":
        low_hits = session_df.loc[prices < low_level, "timestamp"]
    else:
        low_hits = session_df.loc[prices <= low_level, "timestamp"]

    high_ts: pd.Timestamp | None = high_hits.iloc[0] if len(high_hits) else None
    low_ts: pd.Timestamp | None = low_hits.iloc[0] if len(low_hits) else None

    if high_ts is None and low_ts is None:
        sequence = "NO_BREAK"
    elif high_ts is not None and low_ts is None:
        sequence = "HIGH_ONLY"
    elif high_ts is None and low_ts is not None:
        sequence = "LOW_ONLY"
    elif high_ts < low_ts:
        sequence = "HIGH_THEN_LOW"
    elif low_ts < high_ts:
        sequence = "LOW_THEN_HIGH"
    else:
        sequence = "BOTH_SAME_TIME"

    return high_ts, low_ts, sequence


def build_daily_sessions(df: pd.DataFrame, windows: list[SessionWindow], min_trades_per_session: int) -> tuple[pd.DataFrame, dict]:
    print("Building complete Asia/Europe/US daily sessions...")
    df = add_session_day_wib(df)
    dataset_min_ts = df["timestamp"].min()
    dataset_max_ts = df["timestamp"].max()
    records: list[dict] = []
    candidate_days = pd.DatetimeIndex(sorted(df["anchor_utc_date"].dropna().unique()))

    dropped_edge_coverage = 0
    dropped_missing_session = 0
    dropped_min_trades = 0

    for anchor_utc_date in candidate_days:
        windows_map = build_daily_windows(anchor_utc_date, windows)
        asia_start, asia_end = windows_map["asia"]
        europe_start, europe_end = windows_map["europe"]
        us_start, us_end = windows_map["us"]

        # Complete-day check: dataset coverage must span full asia_start -> us_end window.
        if dataset_min_ts > asia_start or dataset_max_ts < us_end:
            dropped_edge_coverage += 1
            continue

        day_record = {
            "session_day_wib": anchor_utc_date.tz_convert(WIB_TZ).strftime("%Y-%m-%d"),
            "anchor_utc_date": anchor_utc_date,
            "asia_start_utc": asia_start,
            "asia_end_utc": asia_end,
            "europe_start_utc": europe_start,
            "europe_end_utc": europe_end,
            "us_start_utc": us_start,
            "us_end_utc": us_end,
        }

        missing_session = False
        for name, (start, end) in windows_map.items():
            mask = (df["timestamp"] >= start) & (df["timestamp"] < end)
            session_df = df.loc[mask].copy()
            if session_df.empty:
                missing_session = True
                break
            if len(session_df) < min_trades_per_session:
                day_record[f"{name}_trade_count"] = int(len(session_df))
                missing_session = False
                day_record["drop_reason"] = "min_trades"
                break

            day_record.update(aggregate_session_metrics(session_df, name))

            if name == "europe":
                high_ts, low_ts, seq = first_break_timestamp(session_df, day_record["asia_high"], day_record["asia_low"], "gt", "lt")
                day_record["europe_first_break_asia_high_timestamp"] = high_ts
                day_record["europe_first_break_asia_low_timestamp"] = low_ts
                day_record["europe_first_break_sequence"] = seq
            elif name == "us":
                pre_us_high = max(day_record["asia_high"], day_record["europe_high"])
                pre_us_low = min(day_record["asia_low"], day_record["europe_low"])
                high_ts, low_ts, seq = first_break_timestamp(session_df, pre_us_high, pre_us_low, "gt", "lt")
                day_record["us_first_break_pre_us_high_timestamp"] = high_ts
                day_record["us_first_break_pre_us_low_timestamp"] = low_ts
                day_record["us_first_break_sequence"] = seq

        if "drop_reason" in day_record and day_record["drop_reason"] == "min_trades":
            dropped_min_trades += 1
            continue
        if missing_session:
            dropped_missing_session += 1
            continue

        records.append(day_record)

    daily = pd.DataFrame(records)
    if daily.empty:
        raise ValueError("No complete session days found after coverage and min-trade filtering.")

    # Human-readable UTC/WIB session columns.
    for prefix in ["asia", "europe", "us"]:
        daily[f"{prefix}_start_wib"] = daily[f"{prefix}_start_utc"].map(format_ts_wib)
        daily[f"{prefix}_end_wib"] = daily[f"{prefix}_end_utc"].map(format_ts_wib)
        daily[f"{prefix}_start_utc_str"] = daily[f"{prefix}_start_utc"].map(format_ts_utc)
        daily[f"{prefix}_end_utc_str"] = daily[f"{prefix}_end_utc"].map(format_ts_utc)

    for col in [
        "europe_first_break_asia_high_timestamp",
        "europe_first_break_asia_low_timestamp",
        "us_first_break_pre_us_high_timestamp",
        "us_first_break_pre_us_low_timestamp",
    ]:
        daily[f"{col}_utc"] = daily[col].map(format_ts_utc)
        daily[f"{col}_wib"] = daily[col].map(format_ts_wib)

    stats = {
        "dataset_min_timestamp_utc": format_ts_utc(dataset_min_ts),
        "dataset_max_timestamp_utc": format_ts_utc(dataset_max_ts),
        "candidate_days": int(len(candidate_days)),
        "complete_days": int(len(daily)),
        "dropped_edge_coverage": int(dropped_edge_coverage),
        "dropped_missing_session": int(dropped_missing_session),
        "dropped_min_trades": int(dropped_min_trades),
    }
    print(f"Complete session days retained: {len(daily):,}")
    return daily, stats


# =============================================================================
# Labeling logic
# =============================================================================


def classify_direction(efficiency: float, threshold: float) -> str:
    if pd.isna(efficiency):
        return "NEUTRAL"
    if efficiency >= threshold:
        return "UP"
    if efficiency <= -threshold:
        return "DOWN"
    return "NEUTRAL"


def classify_body_type(body_to_range: float, low_threshold: float, high_threshold: float) -> str:
    if pd.isna(body_to_range):
        return "UNKNOWN_BODY"
    if body_to_range < low_threshold:
        return "LOW_BODY"
    if body_to_range < high_threshold:
        return "MEDIUM_BODY"
    return "HIGH_BODY"


def apply_range_types(daily: pd.DataFrame, prefix: str, low_q: float, high_q: float) -> pd.DataFrame:
    """
    V1 uses full-sample quantiles for descriptive exploratory research.
    A live/backtest version should switch to expanding or rolling quantiles to avoid look-ahead.
    """
    range_pct_col = f"{prefix}_range_pct"
    low_cut = daily[range_pct_col].quantile(low_q)
    high_cut = daily[range_pct_col].quantile(high_q)

    def classify(x):
        if pd.isna(x):
            return "UNKNOWN_RANGE"
        if x <= low_cut:
            return "LOW_RANGE"
        if x >= high_cut:
            return "HIGH_RANGE"
        return "NORMAL_RANGE"

    daily[f"{prefix}_range_type"] = daily[range_pct_col].map(classify)
    daily[f"{prefix}_range_low_quantile_cut"] = low_cut
    daily[f"{prefix}_range_high_quantile_cut"] = high_cut
    return daily


def add_session_labels(daily: pd.DataFrame, args) -> pd.DataFrame:
    for prefix, label in [("asia", "ASIA"), ("europe", "EUROPE"), ("us", "US")]:
        daily[f"{prefix}_direction"] = daily[f"{prefix}_efficiency"].map(lambda x: classify_direction(x, args.direction_eff_threshold))
        daily[f"{prefix}_body_type"] = daily[f"{prefix}_body_to_range"].map(
            lambda x: classify_body_type(x, args.low_body_threshold, args.high_body_threshold)
        )
        daily[f"{prefix}_condition"] = (
            label
            + "_"
            + daily[f"{prefix}_direction"]
            + "_"
            + daily[f"{prefix}_range_type"]
            + "_"
            + daily[f"{prefix}_body_type"]
        )
    return daily


def add_asia_europe_relationships(daily: pd.DataFrame, args) -> pd.DataFrame:
    daily = daily.copy()
    daily["direction_combo"] = "ASIA_" + daily["asia_direction"] + "_EUROPE_" + daily["europe_direction"]

    daily["europe_asia_range_ratio"] = safe_divide(daily["europe_range"], daily["asia_range"])
    daily["range_dominance"] = np.select(
        [
            daily["europe_asia_range_ratio"] < args.similar_range_low,
            daily["europe_asia_range_ratio"] <= args.similar_range_high,
            daily["europe_asia_range_ratio"] > args.similar_range_high,
        ],
        ["ASIA_DOMINANT", "SIMILAR_RANGE", "EUROPE_DOMINANT"],
        default="SIMILAR_RANGE",
    )

    conditions = [
        (daily["asia_range_type"] == "LOW_RANGE") & (daily["europe_range_type"] == "HIGH_RANGE"),
        (daily["asia_range_type"] == "HIGH_RANGE") & (daily["europe_range_type"] == "LOW_RANGE"),
        (daily["asia_range_type"] == "HIGH_RANGE") & (daily["europe_range_type"] == "HIGH_RANGE"),
        (daily["asia_range_type"] == "LOW_RANGE") & (daily["europe_range_type"] == "LOW_RANGE"),
    ]
    choices = [
        "ASIA_CONTRACTION_EUROPE_EXPANSION",
        "ASIA_EXPANSION_EUROPE_CONTRACTION",
        "BOTH_EXPANSION",
        "BOTH_CONTRACTION",
    ]
    daily["range_sequence"] = np.select(conditions, choices, default="MIXED_NORMAL")

    daily["europe_high_minus_asia_high"] = daily["europe_high"] - daily["asia_high"]
    daily["europe_low_minus_asia_low"] = daily["europe_low"] - daily["asia_low"]

    structure_relation = []
    for _, row in daily.iterrows():
        if row["europe_high"] <= row["asia_high"] and row["europe_low"] >= row["asia_low"]:
            structure_relation.append("EUROPE_INSIDE_ASIA_RANGE")
        elif row["europe_high"] > row["asia_high"] and row["europe_low"] > row["asia_low"]:
            structure_relation.append("EUROPE_HH_HL")
        elif row["europe_high"] < row["asia_high"] and row["europe_low"] < row["asia_low"]:
            structure_relation.append("EUROPE_LH_LL")
        elif row["europe_high"] > row["asia_high"] and row["europe_low"] < row["asia_low"]:
            structure_relation.append("EUROPE_HH_LL")
        else:
            structure_relation.append("EUROPE_PARTIAL_OVERLAP_MIXED")
    daily["structure_relation"] = structure_relation

    overlap = np.maximum(0.0, np.minimum(daily["asia_high"], daily["europe_high"]) - np.maximum(daily["asia_low"], daily["europe_low"]))
    union_range = np.maximum(daily["asia_high"], daily["europe_high"]) - np.minimum(daily["asia_low"], daily["europe_low"])
    daily["overlap"] = overlap
    daily["union_range"] = union_range
    daily["overlap_of_union"] = safe_divide(overlap, union_range)
    daily["overlap_of_asia"] = safe_divide(overlap, daily["asia_range"])
    daily["overlap_of_europe"] = safe_divide(overlap, daily["europe_range"])
    daily["europe_inside_asia_share"] = safe_divide(overlap, daily["europe_range"])
    daily["europe_above_asia_distance"] = np.maximum(0.0, daily["europe_high"] - daily["asia_high"])
    daily["europe_below_asia_distance"] = np.maximum(0.0, daily["asia_low"] - daily["europe_low"])
    daily["europe_above_asia_share"] = safe_divide(daily["europe_above_asia_distance"], daily["europe_range"])
    daily["europe_below_asia_share"] = safe_divide(daily["europe_below_asia_distance"], daily["europe_range"])

    overlap_label = []
    for _, row in daily.iterrows():
        if row["overlap"] == 0 and row["europe_low"] > row["asia_high"]:
            overlap_label.append("NO_OVERLAP_ABOVE")
        elif row["overlap"] == 0 and row["europe_high"] < row["asia_low"]:
            overlap_label.append("NO_OVERLAP_BELOW")
        elif row["overlap_of_union"] >= 0.70:
            overlap_label.append("HIGH_OVERLAP")
        elif row["overlap_of_union"] >= 0.30:
            overlap_label.append("MEDIUM_OVERLAP")
        elif row["overlap_of_union"] > 0:
            overlap_label.append("LOW_OVERLAP")
        else:
            overlap_label.append("NO_OVERLAP_OTHER")
    daily["overlap_label"] = overlap_label
    return daily


def add_europe_behavior(daily: pd.DataFrame, args) -> pd.DataFrame:
    daily = daily.copy()
    daily["combined_high"] = np.maximum(daily["asia_high"], daily["europe_high"])
    daily["combined_low"] = np.minimum(daily["asia_low"], daily["europe_low"])
    daily["combined_range"] = daily["combined_high"] - daily["combined_low"]
    daily["europe_close_position_in_combined_range"] = safe_divide(daily["europe_close"] - daily["combined_low"], daily["combined_range"])

    daily["europe_breaks_asia_high"] = daily["europe_high"] > daily["asia_high"]
    daily["europe_high_break_distance"] = np.maximum(0.0, daily["europe_high"] - daily["asia_high"])
    daily["europe_high_break_share"] = safe_divide(daily["europe_high_break_distance"], daily["combined_range"])

    daily["europe_breaks_asia_low"] = daily["europe_low"] < daily["asia_low"]
    daily["europe_low_break_distance"] = np.maximum(0.0, daily["asia_low"] - daily["europe_low"])
    daily["europe_low_break_share"] = safe_divide(daily["europe_low_break_distance"], daily["combined_range"])

    high_sweep = (
        daily["europe_breaks_asia_high"]
        & (daily["europe_high_break_share"] <= args.sweep_break_share_threshold)
        & (daily["europe_close_position_in_combined_range"] < (1 - args.extreme_close_threshold))
    )
    low_sweep = (
        daily["europe_breaks_asia_low"]
        & (daily["europe_low_break_share"] <= args.sweep_break_share_threshold)
        & (daily["europe_close_position_in_combined_range"] > args.extreme_close_threshold)
    )
    upside_expansion = daily["europe_breaks_asia_high"] & (
        (daily["europe_high_break_share"] > args.sweep_break_share_threshold)
        | (daily["europe_close_position_in_combined_range"] >= (1 - args.extreme_close_threshold))
    )
    downside_expansion = daily["europe_breaks_asia_low"] & (
        (daily["europe_low_break_share"] > args.sweep_break_share_threshold)
        | (daily["europe_close_position_in_combined_range"] <= args.extreme_close_threshold)
    )

    daily["europe_high_sweep"] = high_sweep
    daily["europe_low_sweep"] = low_sweep
    daily["europe_upside_expansion"] = upside_expansion
    daily["europe_downside_expansion"] = downside_expansion

    labels = []
    for _, row in daily.iterrows():
        if not row["europe_breaks_asia_high"] and not row["europe_breaks_asia_low"]:
            labels.append("EUROPE_INSIDE_ASIA_RANGE")
        elif row["europe_high_sweep"] and row["europe_low_sweep"]:
            labels.append("EUROPE_DOUBLE_SWEEP")
        elif row["europe_upside_expansion"] and row["europe_downside_expansion"]:
            labels.append("EUROPE_DOUBLE_EXPANSION")
        elif row["europe_upside_expansion"] and not row["europe_downside_expansion"]:
            labels.append("EUROPE_UPSIDE_EXPANSION")
        elif row["europe_downside_expansion"] and not row["europe_upside_expansion"]:
            labels.append("EUROPE_DOWNSIDE_EXPANSION")
        elif row["europe_high_sweep"] and not row["europe_low_sweep"] and not row["europe_upside_expansion"] and not row["europe_downside_expansion"]:
            labels.append("EUROPE_HIGH_SWEEP")
        elif row["europe_low_sweep"] and not row["europe_high_sweep"] and not row["europe_upside_expansion"] and not row["europe_downside_expansion"]:
            labels.append("EUROPE_LOW_SWEEP")
        elif (row["europe_high_sweep"] or row["europe_low_sweep"]) and (row["europe_upside_expansion"] or row["europe_downside_expansion"]):
            labels.append("EUROPE_MIXED_SWEEP_EXPANSION")
        else:
            labels.append("EUROPE_OTHER")
    daily["europe_behavior_label"] = labels
    return daily


def add_us_outcomes(daily: pd.DataFrame) -> pd.DataFrame:
    daily = daily.copy()
    daily["us_took_asia_high"] = daily["us_high"] > daily["asia_high"]
    daily["us_took_asia_low"] = daily["us_low"] < daily["asia_low"]
    daily["us_took_europe_high"] = daily["us_high"] > daily["europe_high"]
    daily["us_took_europe_low"] = daily["us_low"] < daily["europe_low"]

    daily["pre_us_high"] = np.maximum(daily["asia_high"], daily["europe_high"])
    daily["pre_us_low"] = np.minimum(daily["asia_low"], daily["europe_low"])
    daily["pre_us_range"] = daily["pre_us_high"] - daily["pre_us_low"]

    daily["us_took_pre_us_high"] = daily["us_high"] > daily["pre_us_high"]
    daily["us_took_pre_us_low"] = daily["us_low"] < daily["pre_us_low"]
    daily["us_extension_above_pre_us_high"] = np.maximum(0.0, daily["us_high"] - daily["pre_us_high"])
    daily["us_extension_below_pre_us_low"] = np.maximum(0.0, daily["pre_us_low"] - daily["us_low"])
    daily["us_extension_above_pre_us_high_R"] = safe_divide(daily["us_extension_above_pre_us_high"], daily["pre_us_range"])
    daily["us_extension_below_pre_us_low_R"] = safe_divide(daily["us_extension_below_pre_us_low"], daily["pre_us_range"])
    daily["us_range_vs_asia_range"] = safe_divide(daily["us_range"], daily["asia_range"])
    daily["us_range_vs_europe_range"] = safe_divide(daily["us_range"], daily["europe_range"])
    daily["us_range_vs_pre_us_range"] = safe_divide(daily["us_range"], daily["pre_us_range"])
    daily["us_upside_mfe_from_open"] = safe_divide(daily["us_high"] - daily["us_open"], daily["us_open"])
    daily["us_downside_mfe_from_open"] = safe_divide(daily["us_open"] - daily["us_low"], daily["us_open"])

    daily["us_primary_expansion"] = np.select(
        [daily["us_upside_mfe_from_open"] > daily["us_downside_mfe_from_open"], daily["us_downside_mfe_from_open"] > daily["us_upside_mfe_from_open"]],
        ["UP", "DOWN"],
        default="BALANCED",
    )

    daily["us_behavior_label"] = np.select(
        [
            daily["us_took_pre_us_high"] & ~daily["us_took_pre_us_low"],
            daily["us_took_pre_us_low"] & ~daily["us_took_pre_us_high"],
            daily["us_took_pre_us_high"] & daily["us_took_pre_us_low"],
        ],
        ["US_EXPANDS_UP", "US_EXPANDS_DOWN", "US_EXPANDS_BOTH"],
        default="US_INSIDE_PRE_US_RANGE",
    )

    daily["europe_direction_is_neutral"] = daily["europe_direction"] == "NEUTRAL"
    daily["us_continues_europe_direction"] = np.where(
        daily["europe_direction"] == "UP",
        daily["us_took_europe_high"],
        np.where(daily["europe_direction"] == "DOWN", daily["us_took_europe_low"], np.nan),
    )
    daily["us_reverses_europe_direction"] = np.where(
        daily["europe_direction"] == "UP",
        daily["us_took_europe_low"],
        np.where(daily["europe_direction"] == "DOWN", daily["us_took_europe_high"], np.nan),
    )
    return daily


def build_condition_key(daily: pd.DataFrame) -> pd.DataFrame:
    daily = daily.copy()
    daily["condition_key"] = (
        daily["asia_condition"]
        + "__"
        + daily["europe_condition"]
        + "__"
        + daily["direction_combo"]
        + "__"
        + daily["range_sequence"]
        + "__"
        + daily["range_dominance"]
        + "__"
        + daily["structure_relation"]
        + "__"
        + daily["overlap_label"]
        + "__"
        + daily["europe_behavior_label"]
        + "__"
        + daily["europe_first_break_sequence"]
    )
    return daily


# =============================================================================
# Summaries
# =============================================================================


def summarize_groups(df: pd.DataFrame, group_cols: list[str], include_overlap_medians: bool = False) -> pd.DataFrame:
    grouped = df.groupby(group_cols, dropna=False)
    total_n = len(df)
    rows = []
    for key, g in grouped:
        if not isinstance(key, tuple):
            key = (key,)
        row = {col: val for col, val in zip(group_cols, key)}
        row.update(
            {
                "sample_n": len(g),
                "sample_pct": safe_divide(len(g), total_n),
                "us_expands_up_rate": g["us_behavior_label"].eq("US_EXPANDS_UP").mean(),
                "us_expands_down_rate": g["us_behavior_label"].eq("US_EXPANDS_DOWN").mean(),
                "us_expands_both_rate": g["us_behavior_label"].eq("US_EXPANDS_BOTH").mean(),
                "us_inside_pre_us_range_rate": g["us_behavior_label"].eq("US_INSIDE_PRE_US_RANGE").mean(),
                "us_continues_europe_rate": g["us_continues_europe_direction"].astype("float64").mean(),
                "us_reverses_europe_rate": g["us_reverses_europe_direction"].astype("float64").mean(),
                "median_us_range_pct": g["us_range_pct"].median(),
                "median_us_range_vs_asia_range": g["us_range_vs_asia_range"].median(),
                "median_us_range_vs_europe_range": g["us_range_vs_europe_range"].median(),
                "median_us_range_vs_pre_us_range": g["us_range_vs_pre_us_range"].median(),
                "median_us_extension_above_pre_us_high_R": g["us_extension_above_pre_us_high_R"].median(),
                "median_us_extension_below_pre_us_low_R": g["us_extension_below_pre_us_low_R"].median(),
                "median_us_upside_mfe_from_open": g["us_upside_mfe_from_open"].median(),
                "median_us_downside_mfe_from_open": g["us_downside_mfe_from_open"].median(),
                "median_us_efficiency": g["us_efficiency"].median(),
                "median_us_body_to_range": g["us_body_to_range"].median(),
            }
        )
        if include_overlap_medians:
            row.update(
                {
                    "median_overlap_of_union": g["overlap_of_union"].median(),
                    "median_overlap_of_asia": g["overlap_of_asia"].median(),
                    "median_overlap_of_europe": g["overlap_of_europe"].median(),
                    "median_europe_above_asia_share": g["europe_above_asia_share"].median(),
                    "median_europe_below_asia_share": g["europe_below_asia_share"].median(),
                }
            )
        rows.append(row)
    summary = pd.DataFrame(rows)
    if not summary.empty and "sample_n" in summary.columns:
        summary = summary.sort_values(["sample_n"], ascending=[False]).reset_index(drop=True)
    return summary


def build_top_us_expansion_tables(condition_summary: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    base = condition_summary.loc[condition_summary["sample_n"] >= 10].copy()
    top_up = base.sort_values(["us_expands_up_rate", "sample_n"], ascending=[False, False]).reset_index(drop=True)
    top_down = base.sort_values(["us_expands_down_rate", "sample_n"], ascending=[False, False]).reset_index(drop=True)
    top_range = base.sort_values(["median_us_range_vs_pre_us_range", "sample_n"], ascending=[False, False]).reset_index(drop=True)
    return top_up, top_down, top_range


def sample_manual_validation_rows(df: pd.DataFrame, random_seed: int) -> pd.DataFrame:
    rng = np.random.default_rng(random_seed)
    parts = []

    def sample_per_group(source: pd.DataFrame, group_col: str, n: int, label: str) -> pd.DataFrame:
        rows = []
        for group_value, g in source.groupby(group_col, dropna=False):
            take_n = min(n, len(g))
            chosen_idx = rng.choice(g.index.to_numpy(), size=take_n, replace=False)
            sampled = source.loc[chosen_idx].copy()
            sampled["sample_source"] = label
            sampled["sample_group"] = str(group_value)
            rows.append(sampled)
        return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

    parts.append(sample_per_group(df, "europe_behavior_label", 5, "europe_behavior_label"))
    parts.append(sample_per_group(df, "overlap_label", 5, "overlap_label"))

    major_conditions = df["condition_key"].value_counts()
    major_conditions = major_conditions[major_conditions >= 10].index
    major_df = df[df["condition_key"].isin(major_conditions)].copy()
    if not major_df.empty:
        parts.append(sample_per_group(major_df, "condition_key", 5, "condition_key"))

    out = pd.concat([p for p in parts if not p.empty], ignore_index=True) if any(not p.empty for p in parts) else pd.DataFrame()
    if out.empty:
        return out

    keep_cols = [
        "sample_source",
        "sample_group",
        "session_day_wib",
        "asia_start_utc_str",
        "asia_end_utc_str",
        "europe_start_utc_str",
        "europe_end_utc_str",
        "us_start_utc_str",
        "us_end_utc_str",
        "asia_start_wib",
        "asia_end_wib",
        "europe_start_wib",
        "europe_end_wib",
        "us_start_wib",
        "us_end_wib",
        "condition_key",
        "asia_condition",
        "europe_condition",
        "europe_behavior_label",
        "overlap_label",
        "structure_relation",
        "range_sequence",
        "range_dominance",
        "europe_first_break_sequence",
        "europe_first_break_asia_high_timestamp_utc",
        "europe_first_break_asia_high_timestamp_wib",
        "europe_first_break_asia_low_timestamp_utc",
        "europe_first_break_asia_low_timestamp_wib",
        "us_behavior_label",
        "us_first_break_sequence",
        "us_first_break_pre_us_high_timestamp_utc",
        "us_first_break_pre_us_high_timestamp_wib",
        "us_first_break_pre_us_low_timestamp_utc",
        "us_first_break_pre_us_low_timestamp_wib",
        "asia_open",
        "asia_high",
        "asia_low",
        "asia_close",
        "europe_open",
        "europe_high",
        "europe_low",
        "europe_close",
        "us_open",
        "us_high",
        "us_low",
        "us_close",
        "pre_us_high",
        "pre_us_low",
        "us_extension_above_pre_us_high",
        "us_extension_below_pre_us_low",
        "us_extension_above_pre_us_high_R",
        "us_extension_below_pre_us_low_R",
    ]
    keep_cols = [c for c in keep_cols if c in out.columns]
    return out[keep_cols].sort_values(["sample_source", "sample_group", "session_day_wib"]).reset_index(drop=True)


# =============================================================================
# Excel export
# =============================================================================


PERCENT_KEYWORDS = ["pct", "rate", "share", "position", "_R"]
NUMERIC_KEYWORDS = ["open", "high", "low", "close", "range", "body", "volume", "notional", "distance", "count", "median"]


def auto_adjust_and_format(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, col in enumerate(df.columns, start=1):
        values = [str(col)]
        series = df[col]
        values.extend(["" if pd.isna(v) else str(v) for v in series.head(500)])
        width = min(max(len(v) for v in values) + 2, 40)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in PERCENT_KEYWORDS):
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.number_format = "0.00%"
        elif any(keyword in col_lower for keyword in NUMERIC_KEYWORDS) and pd.api.types.is_numeric_dtype(series):
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.number_format = "0.000000"


def write_top_conditions_sheet(writer: pd.ExcelWriter, sheet_name: str, top_up: pd.DataFrame, top_down: pd.DataFrame, top_range: pd.DataFrame) -> None:
    start_row = 0
    for title, df in [
        ("Top US expands up conditions", top_up),
        ("Top US expands down conditions", top_down),
        ("Top US range vs pre-US range conditions", top_range),
    ]:
        pd.DataFrame({title: []}).to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        start_row += 1
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
        start_row += len(df) + 3

    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        width = min(max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[: min(len(col_cells), 100)]) + 2, 40)
        ws.column_dimensions[letter].width = width


def export_workbook(output_path: Path, config_df: pd.DataFrame, daily: pd.DataFrame, summaries: dict[str, pd.DataFrame], manual_validation: pd.DataFrame) -> None:
    ensure_parent_dir(output_path)
    print(f"Writing Excel workbook: {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        config_df.to_excel(writer, sheet_name="1_config", index=False)
        daily.to_excel(writer, sheet_name="2_daily_sessions", index=False)
        summaries["condition_summary"].to_excel(writer, sheet_name="3_condition_summary", index=False)
        summaries["europe_behavior_summary"].to_excel(writer, sheet_name="4_europe_behavior_summary", index=False)
        summaries["overlap_summary"].to_excel(writer, sheet_name="5_overlap_summary", index=False)
        summaries["direction_combo_summary"].to_excel(writer, sheet_name="6_direction_combo_summary", index=False)
        summaries["range_sequence_summary"].to_excel(writer, sheet_name="7_range_sequence_summary", index=False)
        summaries["structure_summary"].to_excel(writer, sheet_name="8_structure_summary", index=False)

        top_up, top_down, top_range = build_top_us_expansion_tables(summaries["condition_summary"])
        write_top_conditions_sheet(writer, "9_top_us_expansion_conditions", top_up, top_down, top_range)

        manual_validation.to_excel(writer, sheet_name="10_manual_validation_samples", index=False)

        for sheet_name, df in [
            ("1_config", config_df),
            ("2_daily_sessions", daily),
            ("3_condition_summary", summaries["condition_summary"]),
            ("4_europe_behavior_summary", summaries["europe_behavior_summary"]),
            ("5_overlap_summary", summaries["overlap_summary"]),
            ("6_direction_combo_summary", summaries["direction_combo_summary"]),
            ("7_range_sequence_summary", summaries["range_sequence_summary"]),
            ("8_structure_summary", summaries["structure_summary"]),
            ("10_manual_validation_samples", manual_validation),
        ]:
            auto_adjust_and_format(writer, sheet_name, df)
    print("Workbook export complete.")


# =============================================================================
# Main pipeline
# =============================================================================


def build_config_sheet(args, raw_row_count: int, daily_count: int, stats: dict) -> pd.DataFrame:
    items = {
        "run_timestamp_utc": format_ts_utc(pd.Timestamp.now(tz=UTC_TZ)),
        "input_path": args.input,
        "symbol": args.symbol,
        "output_xlsx": args.output_xlsx,
        "raw_row_count": raw_row_count,
        "complete_daily_rows": daily_count,
        "asia_start": args.asia_start,
        "asia_end": args.asia_end,
        "europe_start": args.europe_start,
        "europe_end": args.europe_end,
        "us_start": args.us_start,
        "us_end": args.us_end,
        "direction_eff_threshold": args.direction_eff_threshold,
        "low_body_threshold": args.low_body_threshold,
        "high_body_threshold": args.high_body_threshold,
        "range_low_quantile": args.range_low_quantile,
        "range_high_quantile": args.range_high_quantile,
        "sweep_break_share_threshold": args.sweep_break_share_threshold,
        "extreme_close_threshold": args.extreme_close_threshold,
        "similar_range_low": args.similar_range_low,
        "similar_range_high": args.similar_range_high,
        "min_trades_per_session": args.min_trades_per_session,
        "timezone_output": WIB_TZ,
    }
    items.update(stats)
    return pd.DataFrame({"parameter": list(items.keys()), "value": list(items.values())})


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    validate_args(args)
    windows = build_session_windows(args)

    input_path = Path(args.input)
    output_path = Path(args.output_xlsx)

    df = load_raw_aggtrades(input_path)
    raw_row_count = len(df)

    daily, stats = build_daily_sessions(df, windows, args.min_trades_per_session)

    for prefix in ["asia", "europe", "us"]:
        daily = apply_range_types(daily, prefix, args.range_low_quantile, args.range_high_quantile)

    daily = add_session_labels(daily, args)
    daily = add_asia_europe_relationships(daily, args)
    daily = add_europe_behavior(daily, args)
    daily = add_us_outcomes(daily)
    daily = build_condition_key(daily)

    condition_summary = summarize_groups(daily, ["condition_key"], include_overlap_medians=True)
    europe_behavior_summary = summarize_groups(daily, ["europe_behavior_label"], include_overlap_medians=True)
    overlap_summary = summarize_groups(daily, ["overlap_label"], include_overlap_medians=True)
    direction_combo_summary = summarize_groups(daily, ["direction_combo"], include_overlap_medians=True)
    range_sequence_summary = summarize_groups(daily, ["range_sequence", "range_dominance"], include_overlap_medians=True)
    structure_summary = summarize_groups(daily, ["structure_relation"], include_overlap_medians=True)

    manual_validation = sample_manual_validation_rows(daily, args.random_seed)
    config_df = build_config_sheet(args, raw_row_count, len(daily), stats)

    export_workbook(
        output_path,
        config_df,
        daily,
        {
            "condition_summary": condition_summary,
            "europe_behavior_summary": europe_behavior_summary,
            "overlap_summary": overlap_summary,
            "direction_combo_summary": direction_combo_summary,
            "range_sequence_summary": range_sequence_summary,
            "structure_summary": structure_summary,
        },
        manual_validation,
    )

    print("Done.")
    print(f"Output workbook: {output_path}")


if __name__ == "__main__":
    main()